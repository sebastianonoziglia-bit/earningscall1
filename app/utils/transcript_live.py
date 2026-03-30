"""
Live transcript intelligence — reads directly from the Transcripts Excel sheet.
No pipeline scripts, no SQLite, no CSV files needed.
Works on HuggingFace with only the workbook available.

All keyword lists, weights, and thresholds are imported from scoring_config.py
(single source of truth). Update scoring_config.py to tune vocabulary.
"""
from __future__ import annotations
import re
import logging
import pandas as pd
import streamlit as st

# ── Import all scoring vocabulary from single source of truth ─────────────────
from utils.scoring_config import (
    TOPIC_KEYWORDS, SIGNAL_CATEGORIES, SIGNAL_ICONS, SIGNAL_COLORS,
    OUTLOOK_KEYWORDS, RISK_KEYWORDS, OPPORTUNITY_KEYWORDS,
    INVESTMENT_KEYWORDS, PRODUCT_SHIFT_KEYWORDS, USER_BEHAVIOR_KEYWORDS,
    MONETIZATION_KEYWORDS, STRATEGIC_DIRECTION_KEYWORDS, BROADCASTER_THREAT_KEYWORDS,
    CATEGORY_KEYWORDS,
    FINANCIAL_SCORE_TERMS, FUTURE_TENSE_MARKERS, NEGATION_PREFIXES,
    BOILERPLATE_PHRASES, FORWARD_LOOKING_KEYWORDS,
    CEO_TITLES, CFO_TITLES, KNOWN_ROLE_TITLES, ROLE_NORMALIZER,
    LAYER_WEIGHTS, THRESHOLDS, SPEAKER_ROLE_OVERRIDES,
)

logger = logging.getLogger(__name__)


def score_quote_topics(quote: str) -> list[str]:
    """
    Return which TOPIC_KEYWORDS categories are mentioned in a quote.
    Used to tag signals, filter by topic, and enrich AI context.
    """
    q_lower = quote.lower()
    matched = []
    for topic, keywords in TOPIC_KEYWORDS.items():
        if any(kw in q_lower for kw in keywords):
            matched.append(topic)
    return matched


def enrich_signals_with_topics(signals: list[dict]) -> list[dict]:
    """
    Add a 'topics' list and use it as 'category' if category is missing.
    Call this after extracting any signals/quotes.
    """
    for sig in signals:
        quote = str(sig.get("quote", sig.get("text", ""))).lower()
        topics = score_quote_topics(quote)
        sig["topics"] = topics
        if not sig.get("category") and topics:
            sig["category"] = topics[0]
        elif not sig.get("category"):
            sig["category"] = "General"
    return signals


def _detect_role(title_lower: str) -> str:
    if any(t in title_lower for t in CEO_TITLES):
        return "CEO"
    if any(t in title_lower for t in CFO_TITLES):
        return "CFO"
    return ""


def _score_sentence(sentence: str) -> float:
    s = sentence.lower()
    return sum(1.0 for t in FINANCIAL_SCORE_TERMS if t in s)


def _score_sentence_advanced(
    sentence: str,
    keywords: list,
    role: str,
    sentence_idx: int,
    total_sentences: int,
) -> float:
    """Advanced scorer — all weights from scoring_config.LAYER_WEIGHTS."""
    W = LAYER_WEIGHTS  # shorthand
    s = sentence.strip()
    s_lower = s.lower()

    # 1. Negation check — skip negated sentences
    negation_patterns = [
        "we are not ", "we don't ", "we do not ",
        "we have not ", "we haven't ", "not investing",
        "no longer ", "we stopped ", "we ended ",
    ]
    if any(neg in s_lower for neg in negation_patterns):
        return 0.0

    # 2. Keyword hits — base score
    kw_hits = sum(1 for kw in keywords if kw.lower() in s_lower)
    if kw_hits == 0:
        return 0.0

    # 3. Specificity bonus — sentences with real numbers score higher
    has_number = bool(re.search(
        r'\$[\d,]+|\d+[\.,]\d+[BMK%]|\d{1,3}[BMK]\b|\d+\s*(?:billion|million|percent|%)',
        s, re.IGNORECASE
    ))
    specificity_bonus = W["specificity_bonus"] if has_number else 1.0

    # 4. Forward-looking tense bonus
    forward_phrases = [
        "we will", "we expect", "we are going to",
        "we plan to", "we intend to", "we are targeting",
        "going forward", "next quarter", "next year",
        "in 2025", "in 2026", "we anticipate",
        "we are positioned", "we believe",
    ]
    forward_bonus = W["forward_tense_bonus"] if any(p in s_lower for p in forward_phrases) else 1.0

    # 5. Role bonus
    role_bonus = W["role_bonus_ceo_cfo"] if role in ("CEO", "CFO") else 1.0

    # 6. Position bonus — early in transcript = strategic opening remarks
    position_ratio = 1 - (sentence_idx / max(total_sentences, 1))
    position_bonus = 1.0 + (position_ratio * W["position_max_bonus"])

    # 7. Length factor — prefer medium sentences
    length = len(s)
    if length < 50:
        len_factor = W["len_very_short"]
    elif length < 80:
        len_factor = W["len_short"]
    elif length <= 250:
        len_factor = W["len_medium"]
    else:
        len_factor = W["len_long"]

    # 8. Financial term bonus
    fin_score = sum(W["financial_term_bonus"] for t in FINANCIAL_SCORE_TERMS if t in s_lower)

    base_score = round(
        (kw_hits + fin_score)
        * specificity_bonus
        * forward_bonus
        * role_bonus
        * position_bonus
        * len_factor,
        3,
    )

    # ── Additional verification layers (stacked on base score) ────────────

    # Layer 1 — Future tense detection
    if any(ft in s_lower for ft in FUTURE_TENSE_MARKERS):
        base_score *= W["future_tense_stack"]

    # Layer 2 — Specificity bonus for forward-looking sentences
    # Fires when sentence has concrete numbers AND is forward-looking
    _has_concrete = bool(re.search(
        r'\$[\d,]+[BMbm]?|\d+\.?\d*\s*%|\b20(?:2[4-9]|3[0-9])\b|\bQ[1-4]\b',
        s, re.IGNORECASE,
    ))
    _is_forward = any(ft in s_lower for ft in FUTURE_TENSE_MARKERS)
    if _has_concrete and _is_forward:
        base_score *= W["concrete_forward_stack"]

    # Layer 3 — Negation filter (hard penalty)
    if any(neg in s_lower for neg in NEGATION_PREFIXES):
        base_score *= W["negation_penalty"]

    # Layer 4 — Boilerplate filter
    if any(bp in s_lower for bp in BOILERPLATE_PHRASES):
        base_score *= W["boilerplate_penalty"]

    # Layer 5 — Speaker role bonus already applied above (role_bonus)

    return round(base_score, 3)


def _parse_speaker_blocks(text: str) -> list[dict]:
    """
    Parse transcript text into speaker blocks.
    Handles two formats:
    Format A (dash/colon): "Sundar Pichai -- Chief Executive Officer ..."
    Format B (inline):     "Sundar Pichai Chief Executive Officer Good afternoon..."
    Returns list of {speaker, role, sentences}
    """
    # ── Format A: explicit separator (-- / — / - / :) ──
    speaker_pattern = re.compile(
        r"^([A-Z][A-Za-z\s\.\'\-]{2,50})\s*(?:--|—|-|:)\s*(.{5,100})$"
    )
    # ── Format B: inline role titles (no separator) ──
    # Build regex: "FirstName LastName <known role> <speech...>"
    _role_alts = "|".join(re.escape(r) for r in sorted(KNOWN_ROLE_TITLES, key=len, reverse=True))
    inline_pattern = re.compile(
        r"^([A-Z][A-Za-z\.\'\-]+(?:\s+[A-Z][A-Za-z\.\'\-]+){0,3})\s+"
        r"(" + _role_alts + r")\s+(.{20,})",
        re.DOTALL,
    )

    def _save_block(speaker, role, lines_buf, out):
        if speaker and lines_buf:
            full_text = " ".join(lines_buf)
            sents = [
                s.strip() for s in re.split(r"(?<=[.!?])\s+", full_text)
                if 40 < len(s.strip()) < 500
            ]
            out.append({"speaker": speaker, "role": role, "sentences": sents})

    # First try splitting by double-newline (Format B — inline blocks)
    raw_blocks = re.split(r"\n\s*\n", text)
    blocks: list[dict] = []

    if len(raw_blocks) > 3:
        # Likely Format B (inline blocks separated by blank lines)
        for raw in raw_blocks:
            raw = raw.strip()
            if not raw or len(raw) < 30:
                continue
            if raw.startswith("Company:") or raw.startswith("---"):
                continue
            m = inline_pattern.match(raw)
            if m:
                name = m.group(1).strip()
                role_raw = m.group(2).strip()
                speech = m.group(3).strip()
                role = ROLE_NORMALIZER.get(role_raw, "")
                if not role:
                    role = _detect_role(role_raw.lower())
                if name.lower() in ("operator", "moderator"):
                    role = "Operator"
                # Apply known speaker corrections (overrides transcript-inferred roles)
                _override = SPEAKER_ROLE_OVERRIDES.get(name.lower().strip())
                if _override is not None:
                    role = _override
                sents = [
                    s.strip() for s in re.split(r"(?<=[.!?])\s+", speech)
                    if 40 < len(s.strip()) < 500
                ]
                blocks.append({"speaker": name, "role": role, "sentences": sents})

    if blocks:
        return blocks

    # Fallback: Format A (line-by-line with separators)
    lines = text.split("\n")
    current_speaker = ""
    current_role = ""
    current_lines: list[str] = []

    for line in lines:
        line = line.strip()
        if not line:
            continue
        match = speaker_pattern.match(line)
        if match:
            _save_block(current_speaker, current_role, current_lines, blocks)
            current_lines = []
            name = match.group(1).strip()
            title = match.group(2).strip().lower()
            role = _detect_role(title)
            # Apply known speaker corrections
            _override = SPEAKER_ROLE_OVERRIDES.get(name.lower().strip())
            if _override is not None:
                role = _override
            if role:
                current_speaker = name
                current_role = role
            elif any(w in title for w in ["operator", "analyst", "moderator", "investor"]):
                current_speaker = ""
                current_role = ""
            else:
                current_speaker = name
                current_role = ""
        else:
            if len(line) > 20:
                current_lines.append(line)

    _save_block(current_speaker, current_role, current_lines, blocks)
    return blocks


@st.cache_data(ttl=3600, show_spinner=False)
def extract_ceo_cfo_quotes(
    excel_path: str,
    company: str,
    year: int,
    quarter: str = "",
    max_per_role: int = 3,
) -> dict:
    """
    Extract scored CEO/CFO quotes from Transcripts sheet.
    Returns {"CEO": [{"speaker","role","quote","score"}], "CFO": [...]}
    """
    if not excel_path:
        return {"CEO": [], "CFO": []}
    try:
        df = pd.read_excel(excel_path, sheet_name="Transcripts")
        df.columns = [str(c).strip().lower() for c in df.columns]
        if not {"company", "year", "transcript_text"}.issubset(set(df.columns)):
            return {"CEO": [], "CFO": []}
        df["_c"] = df["company"].astype(str).str.strip().str.lower()
        df["_y"] = pd.to_numeric(df["year"], errors="coerce")
        comp = company.strip().lower()
        mask = (df["_c"] == comp) & (df["_y"] == int(year))
        rows = df[mask]
        if rows.empty:
            rows = df[(df["_c"] == comp) & (df["_y"].between(year - 1, year + 1))]
        if rows.empty:
            return {"CEO": [], "CFO": []}
        if quarter and "quarter" in df.columns:
            q_rows = rows[rows["quarter"].astype(str).str.upper().str.strip() == quarter.upper().strip()]
            if not q_rows.empty:
                rows = q_rows
            # For Annual view: prefer the most recent quarter (sort descending by quarter label)
            elif "quarter" in rows.columns:
                def _q_sort_key(val: str) -> int:
                    s = str(val).upper().strip()
                    for i, qn in enumerate(["Q4", "Q3", "Q2", "Q1"]):
                        if qn in s:
                            return i
                    return 99
                rows = rows.copy()
                rows["_q_sort"] = rows["quarter"].apply(_q_sort_key)
                rows = rows.sort_values("_q_sort")
        text = str(rows.iloc[0].get("transcript_text", "") or "")[:30000]
    except Exception:
        return {"CEO": [], "CFO": []}

    blocks = _parse_speaker_blocks(text)
    result: dict = {"CEO": [], "CFO": []}
    seen_speakers: dict[str, set] = {"CEO": set(), "CFO": set()}
    for block in blocks:
        role = block["role"]
        if role not in result:
            continue
        speaker = block["speaker"]
        # Deduplicate: only show each speaker once per role bucket
        if speaker in seen_speakers[role]:
            continue
        seen_speakers[role].add(speaker)
        scored = sorted(
            [(round(_score_sentence(s), 2), s) for s in block["sentences"] if _score_sentence(s) > 0],
            key=lambda x: -x[0],
        )
        for score, sent in scored[:max_per_role]:
            if len(result[role]) < max_per_role:
                result[role].append({
                    "speaker": speaker,
                    "role": role,
                    "quote": sent,
                    "score": score,
                })
    # Enrich with topic tags
    try:
        for role in result:
            result[role] = enrich_signals_with_topics(result[role])
    except Exception:
        pass
    return result


@st.cache_data(ttl=3600, show_spinner=False)
def extract_pulse_quotes(
    excel_path: str,
    max_quotes: int = 30,
    min_score: float = 2.0,
) -> list[dict]:
    """
    Extract top-scored CEO/CFO quotes across ALL companies for the Human Voice strip.
    Returns list of {company, speaker, role, quote, score, year, quarter}
    """
    if not excel_path:
        return []
    try:
        df = pd.read_excel(excel_path, sheet_name="Transcripts")
        df.columns = [str(c).strip().lower() for c in df.columns]
        if not {"company", "year", "quarter", "transcript_text"}.issubset(set(df.columns)):
            return []
    except Exception:
        return []

    # Get the most recent entry per company
    df["_y"] = pd.to_numeric(df["year"], errors="coerce")
    df = df.dropna(subset=["_y"])
    df["_y"] = df["_y"].astype(int)
    latest_idx = df.groupby("company")["_y"].idxmax()
    df = df.loc[latest_idx].copy()

    results = []
    for _, row in df.iterrows():
        company = str(row.get("company", "")).strip()
        year = int(row.get("_y", 0))
        quarter = str(row.get("quarter", "")).strip()
        text = str(row.get("transcript_text", "") or "")[:30000]
        if not text:
            continue
        blocks = _parse_speaker_blocks(text)
        for block in blocks:
            if block["role"] not in ("CEO", "CFO"):
                continue
            scored = sorted(
                [(round(_score_sentence(s), 2), s) for s in block["sentences"] if _score_sentence(s) >= min_score],
                key=lambda x: -x[0],
            )
            if scored:
                score, sent = scored[0]  # best sentence per speaker block
                results.append({
                    "company": company,
                    "speaker": block["speaker"],
                    "role": block["role"],
                    "quote": sent,
                    "score": score,
                    "year": year,
                    "quarter": quarter,
                })

    results.sort(key=lambda x: -x["score"])
    return results[:max_quotes]


@st.cache_data(ttl=3600, show_spinner=False)
def extract_topic_metrics(
    excel_path: str,
    year: int,
    quarter: str = "",
) -> pd.DataFrame:
    """
    Build topic_metrics DataFrame directly from Transcripts sheet.
    Matches the schema expected by _render_transcript_topic_growth_chart:
    columns: year, quarter, topic, mention_count, companies_mentioned,
             total_companies, importance_pct, growth_pct
    """
    if not excel_path:
        return pd.DataFrame()
    try:
        df = pd.read_excel(excel_path, sheet_name="Transcripts")
        df.columns = [str(c).strip().lower() for c in df.columns]
        if not {"company", "year", "quarter", "transcript_text"}.issubset(set(df.columns)):
            return pd.DataFrame()
    except Exception:
        return pd.DataFrame()

    df["_y"] = pd.to_numeric(df["year"], errors="coerce")
    df = df.dropna(subset=["_y"])
    df["_y"] = df["_y"].astype(int)

    qnum = None
    if quarter:
        m = re.search(r"([1-4])", str(quarter))
        if m:
            qnum = int(m.group(1))

    curr = df[df["_y"] == int(year)]
    if qnum is not None and "quarter" in df.columns:
        q_curr = curr[curr["quarter"].astype(str).str.upper().str.strip() == f"Q{qnum}"]
        if not q_curr.empty:
            curr = q_curr

    if qnum is not None and qnum > 1:
        prior_qnum = qnum - 1
        prior = df[df["_y"] == int(year)]
        prior = prior[prior["quarter"].astype(str).str.upper().str.strip() == f"Q{prior_qnum}"]
    else:
        prior = df[df["_y"] == int(year) - 1]

    total_companies = curr["company"].nunique()
    if total_companies == 0:
        return pd.DataFrame()

    def count_topic(period_df: pd.DataFrame, keywords: list) -> tuple[int, int]:
        mention_count = 0
        companies: set = set()
        for _, row in period_df.iterrows():
            text = str(row.get("transcript_text", "") or "").lower()
            hits = sum(1 for kw in keywords if kw in text)
            if hits > 0:
                mention_count += hits
                companies.add(str(row.get("company", "")).strip())
        return mention_count, len(companies)

    rows = []
    for topic, keywords in TOPIC_KEYWORDS.items():
        curr_count, curr_cos = count_topic(curr, keywords)
        prior_count, _ = count_topic(prior, keywords)
        importance_pct = (curr_cos / total_companies * 100) if total_companies > 0 else 0
        if prior_count > 0:
            growth_pct = (curr_count - prior_count) / prior_count * 100
        elif curr_count > 0:
            growth_pct = 100.0
        else:
            growth_pct = 0.0
        rows.append({
            "year": int(year),
            "quarter": f"Q{qnum}" if qnum else str(year),
            "topic": topic,
            "mention_count": curr_count,
            "companies_mentioned": curr_cos,
            "total_companies": total_companies,
            "importance_pct": round(importance_pct, 2),
            "growth_pct": round(growth_pct, 1),
        })

    result = pd.DataFrame(rows)
    result = result[result["mention_count"] > 0].copy()
    return result


@st.cache_data(ttl=3600, show_spinner=False)
def extract_outlook_risks_opportunities(
    excel_path: str,
    company: str,
    year: int,
    quarter: str = "",
    max_per_category: int | None = None,
) -> dict:
    """
    Extract signals across all 9 categories from CEO/CFO transcript blocks.
    Returns {category: [{speaker, role, quote, score, category}]}
    Uses advanced scoring: negation detection, specificity, forward tense, position.
    """
    if not excel_path:
        return {cat: [] for cat in SIGNAL_CATEGORIES}
    try:
        df = pd.read_excel(excel_path, sheet_name="Transcripts")
        df.columns = [str(c).strip().lower() for c in df.columns]
        if not {"company", "year", "transcript_text"}.issubset(set(df.columns)):
            return {cat: [] for cat in SIGNAL_CATEGORIES}
        df["_c"] = df["company"].astype(str).str.strip().str.lower()
        df["_y"] = pd.to_numeric(df["year"], errors="coerce")
        comp = company.strip().lower()
        mask = (df["_c"] == comp) & (df["_y"] == int(year))
        rows = df[mask]
        if rows.empty:
            rows = df[(df["_c"] == comp) & (df["_y"].between(year - 1, year + 1))]
        if rows.empty:
            return {cat: [] for cat in SIGNAL_CATEGORIES}
        if quarter and "quarter" in df.columns:
            q_rows = rows[rows["quarter"].astype(str).str.upper().str.strip() == quarter.upper().strip()]
            if not q_rows.empty:
                rows = q_rows
            elif "quarter" in rows.columns:
                # Annual view: prefer most recent quarter
                def _q_sort_key2(val: str) -> int:
                    s = str(val).upper().strip()
                    for i, qn in enumerate(["Q4", "Q3", "Q2", "Q1"]):
                        if qn in s:
                            return i
                    return 99
                rows = rows.copy()
                rows["_q_sort"] = rows["quarter"].apply(_q_sort_key2)
                rows = rows.sort_values("_q_sort")
        text = str(rows.iloc[0].get("transcript_text", "") or "")[:30000]
    except Exception:
        return {cat: [] for cat in SIGNAL_CATEGORIES}

    _kw_map = {
        "Outlook":             OUTLOOK_KEYWORDS,
        "Risks":               RISK_KEYWORDS,
        "Opportunities":       OPPORTUNITY_KEYWORDS,
        "Investment":          INVESTMENT_KEYWORDS,
        "Product Shifts":      PRODUCT_SHIFT_KEYWORDS,
        "User Behavior":       USER_BEHAVIOR_KEYWORDS,
        "Monetization":        MONETIZATION_KEYWORDS,
        "Strategic Direction": STRATEGIC_DIRECTION_KEYWORDS,
        "Broadcaster Threats": BROADCASTER_THREAT_KEYWORDS,
    }

    result: dict = {cat: [] for cat in _kw_map}
    seen: dict = {cat: set() for cat in _kw_map}
    blocks = _parse_speaker_blocks(text)

    # Pre-compute total sentence count for position bonus
    _all_sentences = [s for b in blocks for s in b.get("sentences", [])]
    _total_sents = len(_all_sentences)
    _sent_idx = 0

    for block in blocks:
        for sentence in block.get("sentences", []):
            s = sentence.strip()
            _sent_idx += 1
            if len(s) < 40 or len(s) > 400:
                continue
            for category, keywords in _kw_map.items():
                score = _score_sentence_advanced(
                    s, keywords, block.get("role", ""), _sent_idx, _total_sents
                )
                if score == 0.0:
                    continue
                key = s[:60].lower()
                if key in seen[category]:
                    continue
                seen[category].add(key)
                result[category].append({
                    "speaker": block["speaker"],
                    "role": block.get("role", ""),
                    "quote": s,
                    "score": score,
                    "category": category,
                })

    # Category exclusivity — keep sentence in top 2 categories only
    _all_quote_keys: dict = {}
    for cat in result:
        for sig in result[cat]:
            key = sig["quote"][:60].lower()
            if key not in _all_quote_keys:
                _all_quote_keys[key] = []
            _all_quote_keys[key].append((cat, sig["score"]))

    for key, cat_scores in _all_quote_keys.items():
        if len(cat_scores) <= 2:
            continue
        cat_scores.sort(key=lambda x: -x[1])
        keep_cats = {cs[0] for cs in cat_scores[:2]}
        for cat in result:
            if cat not in keep_cats:
                result[cat] = [
                    s for s in result[cat]
                    if s["quote"][:60].lower() != key
                ]

    # Sort by score — best signals first; apply cap only if explicitly set
    for cat in result:
        result[cat] = sorted(result[cat], key=lambda x: -x["score"])
        if max_per_category is not None:
            result[cat] = result[cat][:max_per_category]

    return result


@st.cache_data(ttl=3600, show_spinner=False)
def extract_iconic_quotes(
    excel_path: str,
    year: int,
    quarter: str = "",
    max_quotes: int = 12,
) -> pd.DataFrame:
    """
    Extract best CEO/CFO quotes per company for Overview iconic quotes section.
    Matches schema: year, quarter, company, speaker, role_bucket, quote, score
    """
    pulse = extract_pulse_quotes(excel_path, max_quotes=max_quotes * 3)
    if not pulse:
        return pd.DataFrame()
    rows = []
    seen_companies: set = set()
    for q in pulse:
        co = q["company"]
        if co not in seen_companies:
            seen_companies.add(co)
            rows.append({
                "year": q["year"],
                "quarter": q["quarter"],
                "company": co,
                "speaker": q["speaker"],
                "role_bucket": q["role"],
                "quote": q["quote"],
                "score": q["score"],
            })
        if len(rows) >= max_quotes:
            break
    return pd.DataFrame(rows) if rows else pd.DataFrame()


@st.cache_data(ttl=3600, show_spinner=False)
def extract_forward_looking_signals(
    excel_path: str,
    company: str = "",
    year: int = 0,
    quarter: str = "",
    max_signals: int | None = None,
) -> list[dict]:
    """
    Extract the highest-scoring forward-looking signals for a company.
    Uses multi-layer scoring: future tense + specificity + negation filter.
    Returns list of dicts with quote, speaker, role, score, year, quarter, company, category.
    Used by: Home page CEO carousel, Earnings Forward Intelligence panel,
    Overview cross-company outlook, Genie context building.
    """
    if not excel_path:
        return []
    try:
        df = pd.read_excel(excel_path, sheet_name="Transcripts")
        df.columns = [str(c).strip().lower() for c in df.columns]
        if not {"company", "year", "transcript_text"}.issubset(set(df.columns)):
            return []
    except Exception:
        return []

    df["_c"] = df["company"].astype(str).str.strip().str.lower()
    df["_y"] = pd.to_numeric(df["year"], errors="coerce")
    df = df.dropna(subset=["_y"])
    df["_y"] = df["_y"].astype(int)

    if company:
        comp = company.strip().lower()
        mask = df["_c"] == comp
        rows = df[mask]
        if year:
            yr_rows = rows[rows["_y"] == int(year)]
            if not yr_rows.empty:
                rows = yr_rows
            elif not rows.empty:
                rows = rows[rows["_y"] == rows["_y"].max()]
        if rows.empty:
            return []
        if quarter and "quarter" in df.columns:
            q_rows = rows[rows["quarter"].astype(str).str.upper().str.strip() == quarter.upper().strip()]
            if not q_rows.empty:
                rows = q_rows
    else:
        # Cross-company: use latest year per company
        if year:
            rows = df[df["_y"] == int(year)]
        else:
            latest_idx = df.groupby("_c")["_y"].idxmax()
            rows = df.loc[latest_idx]

    # Combined forward keywords for scoring
    _fwd_kw = FORWARD_LOOKING_KEYWORDS + OUTLOOK_KEYWORDS

    all_signals: list[dict] = []
    seen_keys: set = set()

    for _, row in rows.iterrows():
        _company = str(row.get("company", "")).strip()
        _year = int(row.get("_y", 0))
        _quarter = str(row.get("quarter", "")).strip()
        text = str(row.get("transcript_text", "") or "")[:30000]
        if not text:
            continue

        blocks = _parse_speaker_blocks(text)
        _all_sents = [s for b in blocks for s in b.get("sentences", [])]
        _total = len(_all_sents)
        _idx = 0

        for block in blocks:
            for sentence in block.get("sentences", []):
                s = sentence.strip()
                _idx += 1
                if len(s) < 40 or len(s) > 500:
                    continue
                # Must contain at least one forward-looking keyword
                s_lower = s.lower()
                if not any(kw in s_lower for kw in _fwd_kw):
                    continue

                score = _score_sentence_advanced(
                    s, _fwd_kw, block.get("role", ""), _idx, _total
                )
                if score < 0.5:
                    continue

                key = s[:60].lower()
                if key in seen_keys:
                    continue
                seen_keys.add(key)

                # Determine category from signal keywords
                cat = "Outlook"
                for _cat, _kws in {
                    "Investment": INVESTMENT_KEYWORDS,
                    "Product Shifts": PRODUCT_SHIFT_KEYWORDS,
                    "Opportunities": OPPORTUNITY_KEYWORDS,
                    "Strategic Direction": STRATEGIC_DIRECTION_KEYWORDS,
                }.items():
                    if any(kw in s_lower for kw in _kws):
                        cat = _cat
                        break

                all_signals.append({
                    "quote": s,
                    "speaker": block["speaker"],
                    "role": block.get("role", ""),
                    "score": round(score, 3),
                    "year": _year,
                    "quarter": _quarter,
                    "company": _company,
                    "category": cat,
                    "has_number": bool(re.search(r'\$[\d,]+|\d+\.?\d*\s*%', s)),
                    "has_year_ref": bool(re.search(r'\b20(?:2[4-9]|3[0-9])\b', s)),
                })

    all_signals.sort(key=lambda x: -x["score"])
    result = all_signals[:max_signals] if max_signals is not None else all_signals

    # Enrich with real speaker names from Company_Speakers sheet
    for sig in result:
        _sp = sig.get("speaker", "")
        _rl = sig.get("role", "")
        if _rl in ("CEO", "CFO", "COO", "CTO", "CRO", "CMO") and excel_path:
            _full = get_speaker_name(sig["company"], _rl, excel_path)
            if _full:
                sig["speaker"] = _full
                sig["speaker_display"] = f"{_full} \u00b7 {_rl}"
            else:
                sig["speaker_display"] = f"{_sp} \u00b7 {_rl}" if _rl else _sp
        else:
            sig["speaker_display"] = f"{_sp} \u00b7 {_rl}" if _rl else _sp

    return result


def extract_forward_looking_signals_batch(
    excel_path: str,
    companies: list[str],
    year: int = 0,
    max_signals_per_company: int | None = None,
) -> dict[str, list[dict]]:
    """Batch version: reads Transcripts sheet ONCE, returns signals for all companies.

    Returns dict mapping company name → list of signal dicts.
    """
    if not excel_path or not companies:
        return {}
    try:
        df = pd.read_excel(excel_path, sheet_name="Transcripts")
        df.columns = [str(c).strip().lower() for c in df.columns]
        if not {"company", "year", "transcript_text"}.issubset(set(df.columns)):
            return {}
    except Exception:
        return {}

    df["_c"] = df["company"].astype(str).str.strip().str.lower()
    df["_y"] = pd.to_numeric(df["year"], errors="coerce")
    df = df.dropna(subset=["_y"])
    df["_y"] = df["_y"].astype(int)

    _fwd_kw = FORWARD_LOOKING_KEYWORDS + OUTLOOK_KEYWORDS
    result: dict[str, list[dict]] = {}

    for company in companies:
        comp = company.strip().lower()
        mask = df["_c"] == comp
        rows = df[mask]
        if year:
            yr_rows = rows[rows["_y"] == int(year)]
            if not yr_rows.empty:
                rows = yr_rows
            elif not rows.empty:
                rows = rows[rows["_y"] == rows["_y"].max()]
        if rows.empty:
            continue

        all_signals: list[dict] = []
        seen_keys: set = set()

        for _, row in rows.iterrows():
            _company = str(row.get("company", "")).strip()
            _year = int(row.get("_y", 0))
            _quarter = str(row.get("quarter", "")).strip()
            text = str(row.get("transcript_text", "") or "")[:30000]
            if not text:
                continue

            blocks = _parse_speaker_blocks(text)
            _all_sents = [s for b in blocks for s in b.get("sentences", [])]
            _total = len(_all_sents)
            _idx = 0

            for block in blocks:
                for sentence in block.get("sentences", []):
                    s = sentence.strip()
                    _idx += 1
                    if len(s) < 40 or len(s) > 500:
                        continue
                    s_lower = s.lower()
                    if not any(kw in s_lower for kw in _fwd_kw):
                        continue

                    score = _score_sentence_advanced(
                        s, _fwd_kw, block.get("role", ""), _idx, _total
                    )
                    if score < 0.5:
                        continue

                    key = s[:60].lower()
                    if key in seen_keys:
                        continue
                    seen_keys.add(key)

                    cat = "Outlook"
                    for _cat, _kws in {
                        "Investment": INVESTMENT_KEYWORDS,
                        "Product Shifts": PRODUCT_SHIFT_KEYWORDS,
                        "Opportunities": OPPORTUNITY_KEYWORDS,
                        "Strategic Direction": STRATEGIC_DIRECTION_KEYWORDS,
                    }.items():
                        if any(kw in s_lower for kw in _kws):
                            cat = _cat
                            break

                    all_signals.append({
                        "quote": s,
                        "speaker": block["speaker"],
                        "role": block.get("role", ""),
                        "score": round(score, 3),
                        "year": _year,
                        "quarter": _quarter,
                        "company": _company,
                        "category": cat,
                    })

        all_signals.sort(key=lambda x: -x["score"])
        signals = all_signals[:max_signals_per_company] if max_signals_per_company else all_signals

        for sig in signals:
            _sp = sig.get("speaker", "")
            _rl = sig.get("role", "")
            if _rl in ("CEO", "CFO", "COO", "CTO", "CRO", "CMO") and excel_path:
                _full = get_speaker_name(sig["company"], _rl, excel_path)
                if _full:
                    sig["speaker"] = _full
                    sig["speaker_display"] = f"{_full} \u00b7 {_rl}"
                else:
                    sig["speaker_display"] = f"{_sp} \u00b7 {_rl}" if _rl else _sp
            else:
                sig["speaker_display"] = f"{_sp} \u00b7 {_rl}" if _rl else _sp

        if signals:
            result[company] = signals

    return result


# TODO: extract_forward_looking_signals could also power:
# - Automated quarterly briefing generation (one-click "Q4 2024 briefing")
# - Editorial page forward-looking content (currently uses simpler keyword matching)
# - Overview cross-company forward intelligence cards
# - Automated Genie scenario context for "what if" questions


@st.cache_data(ttl=3600, show_spinner=False)
def extract_all_signals(
    excel_path: str,
    company: str,
    year: int,
    quarter: str = "",
) -> list[dict]:
    """
    Extract signals across ALL categories for a company.
    Wrapper around extract_outlook_risks_opportunities that returns flat list.
    Used by Overview Transcript Signal Explorer.
    """
    result = extract_outlook_risks_opportunities(
        excel_path, company, year, quarter,
    )
    flat: list[dict] = []
    for cat, signals in result.items():
        for sig in signals:
            sig["company"] = company
            sig["year"] = year
            sig["quarter"] = quarter
            flat.append(sig)
    return flat


# ══════════════════════════════════════════════════════════════════════════════
# SPEAKER REGISTRY — extract, persist, and look up real speaker names
# ══════════════════════════════════════════════════════════════════════════════

def extract_speakers_from_transcript(
    transcript_text: str,
    company: str,
    year: int,
    quarter: str,
) -> list[dict]:
    """
    Parse transcript text to extract speaker names and roles.
    Format: "FirstName LastName RoleTitle Speech..."
    Returns list of dicts with company, year, quarter, full_name, role_raw,
    role_normalized, is_executive. Deduplicates per company.
    """
    speakers: list[dict] = []
    seen_names: set = set()

    raw_blocks = re.split(r"\n\s*\n", transcript_text)

    for block in raw_blocks:
        block = block.strip()
        if not block or len(block) < 20:
            continue
        if block.startswith("Company:") or block.startswith("---"):
            continue

        for role_title in sorted(KNOWN_ROLE_TITLES, key=len, reverse=True):
            if role_title in block:
                idx = block.index(role_title)
                name_part = block[:idx].strip()
                name_words = name_part.split()
                if 1 <= len(name_words) <= 4:
                    full_name = " ".join(name_words)
                    if full_name.lower() in ("operator", "moderator", ""):
                        break
                    key = f"{company}::{full_name.lower()}"
                    if key in seen_names:
                        break
                    seen_names.add(key)

                    role_normalized = ROLE_NORMALIZER.get(role_title, role_title)
                    is_executive = role_normalized in (
                        "CEO", "CFO", "COO", "CTO", "President",
                        "SVP", "EVP", "VP", "CRO", "CMO", "CPO",
                    )
                    speakers.append({
                        "company": company,
                        "year": int(year),
                        "quarter": quarter,
                        "full_name": full_name,
                        "role_raw": role_title,
                        "role_normalized": role_normalized,
                        "is_executive": is_executive,
                    })
                break  # only match first role per block

    return speakers


def build_speaker_registry(excel_path: str) -> list[dict]:
    """
    Scan ALL transcripts and build a deduplicated speaker registry.
    Each speaker appears once per company (most recent year wins).
    """
    try:
        df = pd.read_excel(excel_path, sheet_name="Transcripts")
        df.columns = [str(c).strip().lower() for c in df.columns]
    except Exception:
        return []

    all_speakers: dict = {}
    for _, row in df.iterrows():
        company = str(row.get("company", "")).strip()
        year = pd.to_numeric(pd.Series([row.get("year")]), errors="coerce").iloc[0]
        quarter = str(row.get("quarter", "")).strip()
        txt = str(row.get("transcript_text", "") or "")
        if not txt or not company or pd.isna(year):
            continue
        for sp in extract_speakers_from_transcript(txt, company, int(year), quarter):
            key = f"{company}::{sp['full_name'].lower()}"
            if key not in all_speakers or sp["year"] > all_speakers[key]["year"]:
                all_speakers[key] = sp

    return list(all_speakers.values())


def write_speakers_to_excel(excel_path: str) -> int:
    """
    Build speaker registry and write Company_Speakers sheet to workbook.
    Returns number of speakers written.
    """
    from openpyxl import load_workbook

    speakers = build_speaker_registry(excel_path)
    if not speakers:
        return 0

    df = pd.DataFrame(speakers)
    df = df.sort_values(["company", "role_normalized", "full_name"])
    df = df[["company", "full_name", "role_normalized", "role_raw",
             "is_executive", "year", "quarter"]]
    df.columns = ["Company", "Full Name", "Role", "Role (raw)",
                  "Is Executive", "Last Seen Year", "Last Seen Quarter"]

    wb = load_workbook(excel_path)
    if "Company_Speakers" in wb.sheetnames:
        del wb["Company_Speakers"]
    ws = wb.create_sheet("Company_Speakers")
    ws.append(list(df.columns))
    for _, row in df.iterrows():
        ws.append(list(row))
    wb.save(excel_path)
    return len(df)


@st.cache_data(ttl=3600, show_spinner=False)
def get_speaker_name(company: str, role: str, excel_path: str) -> str | None:
    """Look up a speaker's full name by company and role from Company_Speakers sheet."""
    try:
        df = pd.read_excel(excel_path, sheet_name="Company_Speakers")
        df.columns = [str(c).strip() for c in df.columns]
        match = df[
            (df["Company"].str.lower() == company.lower())
            & (df["Role"].str.upper() == role.upper())
        ]
        if not match.empty:
            return str(match.iloc[0]["Full Name"])
    except Exception:
        pass
    return None
