from __future__ import annotations

from concurrent.futures import ThreadPoolExecutor, TimeoutError as FutureTimeoutError
from dataclasses import dataclass
from datetime import datetime, timezone
import logging
from pathlib import Path
import re
from typing import Dict, Iterable, List, Optional, Set, Tuple

from openpyxl import load_workbook
from utils.workbook_source import resolve_financial_data_xlsx


TRANSCRIPT_FILE_RE = re.compile(r"^Q([1-4])\.txt$", re.IGNORECASE)
MAX_XLSX_CELL_CHARS = 32767
logger = logging.getLogger(__name__)


@dataclass
class SyncResult:
    scanned_files: int = 0
    existing_rows: int = 0
    appended_rows: int = 0
    truncated_rows: int = 0
    workbook_path: str = ""
    error: str = ""


def _normalize_company(value: str) -> str:
    return re.sub(r"\s+", " ", str(value or "").replace("_", " ").strip())


def _normalize_quarter(value: str) -> str:
    text = str(value or "").strip().upper()
    if text.startswith("Q") and len(text) >= 2 and text[1].isdigit():
        q = int(text[1])
        if 1 <= q <= 4:
            return f"Q{q}"
    if text.isdigit():
        q = int(text)
        if 1 <= q <= 4:
            return f"Q{q}"
    match = re.search(r"\b([1-4])\b", text)
    if match:
        return f"Q{int(match.group(1))}"
    return text


def _key(company: str, year: int, quarter: str) -> str:
    return f"{_normalize_company(company).lower()}_{int(year)}_{_normalize_quarter(quarter)}"


def _resolve_local_workbook_path() -> Optional[Path]:
    resolved = resolve_financial_data_xlsx([])
    if not resolved:
        return None
    path = Path(resolved)
    return path.resolve() if path.exists() else None


def _collect_transcript_files(transcript_root: Path) -> List[Tuple[str, int, str, Path]]:
    rows: List[Tuple[str, int, str, Path]] = []
    if not transcript_root.exists():
        return rows

    for company_dir in sorted([p for p in transcript_root.iterdir() if p.is_dir()]):
        company = _normalize_company(company_dir.name)
        for year_dir in sorted([p for p in company_dir.iterdir() if p.is_dir()]):
            if not year_dir.name.isdigit():
                continue
            year = int(year_dir.name)
            for txt_path in sorted([p for p in year_dir.iterdir() if p.is_file() and p.suffix.lower() == ".txt"]):
                match = TRANSCRIPT_FILE_RE.match(txt_path.name)
                if not match:
                    continue
                quarter = f"Q{int(match.group(1))}"
                rows.append((company, year, quarter, txt_path))
    return rows


def _column_index_map(header_cells: Iterable) -> Dict[str, int]:
    mapping: Dict[str, int] = {}
    for idx, cell in enumerate(header_cells, start=1):
        key = re.sub(r"[^a-z0-9]+", "", str(cell or "").strip().lower())
        if key:
            mapping[key] = idx
    return mapping


def _sheet_key_set(ws) -> Tuple[Set[str], int]:
    max_col = max(ws.max_column, 5)
    rows_iter = ws.iter_rows(min_row=1, max_row=1, min_col=1, max_col=max_col, values_only=True)
    header = next(rows_iter, None)

    # Header must include the canonical fields. If not, rebuild from fixed column positions.
    col_map = _column_index_map(header or [])
    company_col = col_map.get("company", 1)
    year_col = col_map.get("year", 2)
    quarter_col = col_map.get("quarter", 3)

    keys: Set[str] = set()
    existing_rows = 0
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=1, max_col=max_col, values_only=True):
        company = row[company_col - 1] if company_col - 1 < len(row) else None
        year = row[year_col - 1] if year_col - 1 < len(row) else None
        quarter = row[quarter_col - 1] if quarter_col - 1 < len(row) else None
        if company in (None, "") or year in (None, "") or quarter in (None, ""):
            continue
        try:
            year_int = int(float(year))
        except Exception:
            continue
        keys.add(_key(str(company), year_int, str(quarter)))
        existing_rows += 1
    return keys, existing_rows


def _sync_local_transcripts_to_workbook_impl() -> SyncResult:
    result = SyncResult()
    workbook_path = _resolve_local_workbook_path()
    if workbook_path is None:
        result.error = "Local workbook not found."
        return result
    result.workbook_path = str(workbook_path)

    repo_root = Path(__file__).resolve().parents[2]
    transcript_root = repo_root / "earningscall_transcripts"
    transcript_files = _collect_transcript_files(transcript_root)
    result.scanned_files = len(transcript_files)
    if not transcript_files:
        return result

    wb = load_workbook(workbook_path)
    if "Transcripts" in wb.sheetnames:
        ws = wb["Transcripts"]
    else:
        ws = wb.create_sheet("Transcripts")
        ws.append(["company", "year", "quarter", "transcript_text", "last_updated"])

    existing_keys, existing_rows = _sheet_key_set(ws)
    result.existing_rows = existing_rows

    last_updated = datetime.now(timezone.utc).strftime("%Y-%m-%d %H:%M:%S UTC")
    appended = 0
    truncated = 0

    for company, year, quarter, txt_path in transcript_files:
        k = _key(company, year, quarter)
        if k in existing_keys:
            continue
        text = txt_path.read_text(encoding="utf-8", errors="ignore").strip()
        if len(text) > MAX_XLSX_CELL_CHARS:
            text = text[:MAX_XLSX_CELL_CHARS]
            truncated += 1
        ws.append([company, int(year), quarter, text, last_updated])
        existing_keys.add(k)
        appended += 1

    if appended > 0:
        wb.save(workbook_path)

    result.appended_rows = appended
    result.truncated_rows = truncated
    return result


def sync_local_transcripts_to_workbook(timeout_seconds: int = 10) -> SyncResult:
    """Best-effort startup sync that never raises to callers.

    Runs with a hard wait timeout so app startup can continue even when
    workbook/network dependencies are unavailable.
    """
    timeout_seconds = max(int(timeout_seconds or 10), 1)
    executor = ThreadPoolExecutor(max_workers=1, thread_name_prefix="startup_transcript_sync")
    future = executor.submit(_sync_local_transcripts_to_workbook_impl)
    try:
        return future.result(timeout=timeout_seconds)
    except FutureTimeoutError:
        logger.warning(
            "Startup transcript sync timed out after %ss; continuing without transcript sync.",
            timeout_seconds,
        )
        future.cancel()
        return SyncResult(error=f"Timed out after {timeout_seconds}s")
    except Exception as exc:
        logger.warning(
            "Startup transcript sync failed; continuing without transcript sync: %s",
            exc,
            exc_info=True,
        )
        return SyncResult(error=str(exc))
    finally:
        executor.shutdown(wait=False, cancel_futures=True)
