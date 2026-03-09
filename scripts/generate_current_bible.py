#!/usr/bin/env python3
from __future__ import annotations

import ast
import datetime as dt
import json
import re
import subprocess
import time
from collections import defaultdict
from dataclasses import dataclass
from pathlib import Path
from typing import Iterable

ROOT = Path(__file__).resolve().parents[1]
OUT = ROOT / "reports" / "Developer_Insights_Bible_CURRENT.md"

KEY_FILES = [
    "app.py",
    "app/Welcome.py",
    "app/pages/00_Overview.py",
    "app/pages/01_Earnings.py",
    "app/pages/02_Stocks.py",
    "app/pages/03_Editorial.py",
    "app/pages/04_Genie.py",
    "app/data_processor.py",
    "app/stock_processor_fix.py",
    "app/utils/workbook_source.py",
    "app/utils/workbook_market_data.py",
    "app/utils/transcript_startup_sync.py",
    "scripts/rebuild_transcript_index.py",
    "scripts/extract_transcript_topics.py",
    "scripts/extract_transcript_highlights_from_sheet.py",
    "scripts/extract_kpi_values.py",
    "scripts/build_intelligence_db.py",
    "scripts/sync_gsheet_to_sql.py",
    "scripts/generate_insights.py",
    "scripts/sync_all_intelligence.py",
]

# Human-readable description of every page in the app
PAGE_DESCRIPTIONS = {
    "00_Overview.py": "Macro overview — global ad market, duopoly analysis, M2 money supply, concentration bars, landscape bubble chart, market bet. 8-section navigator with per-section st.stop() exits.",
    "01_Earnings.py": "Company deep-dive — hero banner, KPI cards, revenue waterfall, segment composition donut, segment evolution chart, transcript highlights, heatmap, AI chat (Genie).",
    "02_Stocks.py": "Stock price charts, returns, market cap trends, volume, daily/minute OHLCV data sourced from the 'Stocks & Crypto' sheet.",
    "03_Editorial.py": "Editorial/insight narrative page. Renders auto-generated company and macro insights from the intelligence pipeline.",
    "04_Genie.py": "AI assistant (Genie) — natural-language Q&A about earnings data, powered by OpenAI API with SQLite transcript intelligence DB as context.",
}

# Canonical sheet inventory: name → (purpose, key columns, used by)
SHEET_CATALOG = [
    ("Stocks & Crypto",             "Daily OHLCV price + market cap for all tracked tickers",          "date, price, open, high, low, vol., change%, market cap",          "02_Stocks.py, Welcome.py (market tape)"),
    ("USD Inflation",               "Historical US CPI / inflation rates (official + shadow)",          "Year, Official Headline CPI, Core Inflation, ShadowStats*",         "00_Overview.py (macro panel)"),
    ("Nasdaq Composite Est. (FRED)","Nasdaq index + estimated market cap time-series (FRED)",           "observation_date, NASDAQCOM, estimated_nasdaq_market_cap_usd",      "00_Overview.py"),
    ("Country_Totals_vs_GDP",       "Ad spend per country vs GDP (yearly, country-level)",              "Country, Year, AdSpending_Total, AdSpending_USD, GDP_USD, Ad_vs_GDP_%", "Welcome.py (globe choropleth)"),
    ("Country_Advertising_Data_FullVi", "Full ad spend breakdown by country/year/metric_type",        "Country, Year, Metric_type, Value",                                 "00_Overview.py"),
    ("Country_avg_timespent_intrnt24",  "Average daily internet time per country (2024)",              "Country, Daily Time Spent Internet (hours.minutes)",                "00_Overview.py"),
    ("Global_Adv_Aggregates",       "Global ad market totals by year and metric_type",                 "year, metric_type, value",                                          "Welcome.py, 00_Overview.py"),
    ("Global Advertising (GroupM)", "GroupM annual global ad spend by format (TV, Search, etc.)",      "Year, Traditional_TV, Connected_TV, Search, NonSearch, Retail_Media","Welcome.py hero KPI, 00_Overview.py"),
    (" (GroupM) Granular ",         "GroupM granular ad spend with more format detail",                "Year, TV/Pro Video, Audio, Newspapers, Internet, ...",              "00_Overview.py"),
    ("M2_values",                   "US M2 money supply (legacy sheet name; new sheet is 'M2')",       "observation_date / USD observation_date, WM2NS / M2SL",             "00_Overview.py, 01_Earnings.py, utils/m2_supply_data.py"),
    ("Company_metrics_earnings_values", "Annual financial metrics per company (the main data table)", "Company, Year, Revenue, Operating Income, Net Income, Debt, R&D, CapEx, EPS, FCF, EBITDA, Margin", "01_Earnings.py, 00_Overview.py, data_processor.py"),
    ("Company_Employees",           "Employee headcount per company per year",                         "Company, Year, Employee Count",                                     "01_Earnings.py (pie chart center label)"),
    ("Company_Segments_insights_text","AI-generated segment-level insights per company/year/quarter",  "Company, Year, Segment, Insight, Category",                         "01_Earnings.py (segment insights panel)"),
    ("Company_insights_text",       "AI-generated company-level insights (annual + quarterly)",        "company, year, category, insight, quarter",                         "01_Earnings.py (insights panel)"),
    ("Company_advertising_revenue", "Advertising revenue per company per year",                        "Year, Google_Ads, Meta_Ads, Amazon_Ads, Spotify_Ads, ...",          "Welcome.py (revenue anatomy), 00_Overview.py"),
    ("Company_subscribers_values",  "Subscriber counts per service/quarter/year",                      "service, quarter, year, subscribers, unit, US_Canada, International","01_Earnings.py"),
    ("Hardware_Smartphone_Shipments","Global smartphone shipment data by OEM and year",                "Year, Apple_iPhone_Units_M, Samsung_Units_M, ...",                  "00_Overview.py"),
    ("Macro_Wealth_by_Generation",  "Generational wealth distribution data by country/year",           "Country, Year, Age_Group, Generation_Label, Total_Wealth_Billion_USD, Wealth_Share_Pct", "00_Overview.py"),
    ("Company_revenue_by_region",   "Revenue split by geographic region per company/year",             "company, year, segment_name, revenue_millions",                     "01_Earnings.py"),
    ("Company_minute&dollar_earned","Revenue-per-minute-watched and engagement metrics",               "Platform, Subscribers/Users, Avg Time per User, Total Minutes, Revenue, $ per Minute", "Welcome.py (dependency chart / landscape)"),
    ("Company_yearly_segments_values","Annual segment revenue per company",                            "Company, year, segments, Yearly Segment Revenue",                   "01_Earnings.py (segment composition + evolution)"),
    ("Company_Quarterly_segments_valu","Quarterly financial metrics per company (income statement)",   "Ticker, Year, Revenue, Cost Of Revenue, Operating Income, Net Income, Capex, R&D, ...", "01_Earnings.py"),
    ("Alphabet Quarterly Segments", "Alphabet segment revenue by quarter",                             "Quarter, Search & Other, YouTube Ads, Google Network, Google Cloud, ...", "01_Earnings.py"),
    ("Apple Quarterly Segments",    "Apple segment revenue by quarter",                                "Quarter, iPhone, Mac, iPad, Wearables, Services",                   "01_Earnings.py"),
    ("Amazon Quarterly Segments ",  "Amazon segment revenue by quarter",                               "Quarter, Online Stores, 3P Seller, AWS, Advertising, ...",          "01_Earnings.py"),
    ("Meta Quarterly Segments",     "Meta segment revenue by quarter",                                 "Quarter, Family of Apps, Reality Labs",                             "01_Earnings.py"),
    ("Comcast Quarterly Segments Gran","Comcast granular quarterly segment breakdown",                  "Quarter, Domestic Broadband, Wireless, Video, Advertising, ...",    "01_Earnings.py"),
    ("Disney Quarterly Segments",   "Disney segment revenue by quarter",                               "Quarter, ...",                                                      "01_Earnings.py"),
    ("Microsoft Quarterly Segments","Microsoft segment revenue by quarter",                            "Quarter, Cloud, Office, Windows, Gaming, LinkedIn, Search, ...",    "01_Earnings.py"),
    ("Netflix Quarterly Segments",  "Netflix revenue by region per quarter",                           "Quarter, UCAN, EMEA, LATAM, APAC",                                  "01_Earnings.py"),
    ("Paramount Quarterly Segments","Paramount segment revenue by quarter",                            "Quarter, DTC, TV Media, Filmed Entertainment",                      "01_Earnings.py"),
    ("Roku Quarterly Segments",     "Roku segment revenue by quarter",                                 "Quarter, Platform, Devices/Player",                                 "01_Earnings.py"),
    ("Spotify Quarterly Segments",  "Spotify segment revenue by quarter",                              "Quarter, Premium, Ad-supported",                                    "01_Earnings.py"),
    ("Warner Bros Quarterly Segments","WBD segment revenue by quarter",                                "Quarter, Distribution, Advertising, Content, Other",                "01_Earnings.py"),
    ("Overview_Macro",              "Curated macro snapshot row for Overview dashboard",               "year, quarter, m2_value, global_ad_market, duopoly_share, ...",     "00_Overview.py"),
    ("Overview_Insights",           "Manually curated insight rows for Overview page",                 "insight_id, sort_order, category, title, year, quarter, comment",  "00_Overview.py"),
    ("Overview_Charts",             "Chart metadata (titles, comments) for Overview charts",           "chart_key, year, quarter, title, pre_comment, post_comment",        "00_Overview.py"),
    ("Transcripts",                 "Raw earnings call transcript text per company/year/quarter",       "company, year, quarter, transcript_text, last_updated",             "scripts/rebuild_transcript_index.py, intelligence pipeline"),
    ("Overview_Auto_Insights",      "Auto-generated macro insights for Overview page",                 "insight_id, category, title, text, comment, priority, companies",  "00_Overview.py"),
]

# Utils catalog: filename → one-line description
UTILS_CATALOG = {
    "workbook_source.py":    "Resolves the primary Excel data file — tries local files first (attached_assets/*.xlsx), then falls back to Google Sheets download with 20s timeout. Validates sheet count, required tabs, and financial coverage.",
    "workbook_market_data.py": "Loads stock/market data from the 'Stocks & Crypto' sheet. Builds company-ticker maps for market tape and 02_Stocks.py.",
    "state_management.py":  "Provides get_data_processor() — a cached singleton FinancialDataProcessor instance shared across all pages via Streamlit session state.",
    "data_granularity.py":  "Detects available data granularity (annual/quarterly/monthly) per company from the workbook sheets.",
    "data_availability.py": "Returns available quarters for a given company/year pair from quarterly segment sheets.",
    "m2_supply_data.py":    "Loads M2 money supply data. Supports both old sheet ('M2_values', columns WM2NS/USD observation_date) and new sheet ('M2', columns M2SL/observation_date) with automatic fallback.",
    "data_loader.py":       "Generic data loading helpers used by data_processor.py.",
    "auth.py":              "Password gate (check_password()). Called on pages that require authentication.",
    "header.py":            "Renders the top navigation header (display_header()). Shared across all pages.",
    "logos.py":             "Loads company logos from attached_assets/ as base64. Used in market tape, Earnings hero, pulse strip.",
    "global_fonts.py":      "Injects Google Fonts (DM Sans, Syne, Montserrat) via st.markdown. Called at top of every page.",
    "styles.py":            "CSS + Plotly theme definitions. get_page_style() returns the light-theme CSS used by Earnings/Overview. apply_plotly_theme() sets the default Plotly template ('mfe_blue' based on plotly_white).",
    "theme.py":             "get_theme_mode() — returns current light/dark preference.",
    "components.py":        "render_ai_assistant() — renders the Genie AI chat UI component.",
    "transcript_startup_sync.py": "One-time-per-container sync of local transcript text files into the Excel workbook's 'Transcripts' sheet.",
    "live_stock_feed.py":   "Fetches live or cached stock prices for the market tape strip on Welcome.py.",
    "macro_trends.py":      "Macro data helpers (Fed Funds rate, inflation, M2 trend calculations).",
    "fed_funds_data.py":    "Loads Federal Funds Rate data for macro panel.",
    "inflation_analysis.py":"Inflation analysis helpers — CPI comparisons, real vs nominal calculations.",
    "insights.py":          "Generates/retrieves insight text for display on Overview and Earnings pages.",
    "insights_loader.py":   "Loads generated_insights_latest.csv for display.",
    "database_service.py":  "SQLite service wrapper for earningscall_intelligence.db.",
    "api_client.py":        "API client for external data sources.",
    "genie_ai.py":          "Core Genie AI logic — builds context from transcript DB + sends to OpenAI.",
    "openai_service.py":    "OpenAI API wrapper used by Genie.",
    "ai_assistant.py":      "AI assistant UI helpers.",
    "page_transition.py":   "Page transition animation helpers.",
    "helpers.py":           "General utility functions shared across pages.",
}

# Environment variables used by the app
ENV_VARS = [
    ("GOOGLE_SHEET_URL",    "Required", "URL to the Google Sheet XLSX export. Used by workbook_source.py as fallback if no local XLSX found."),
    ("AUTO_REFRESH_INTELLIGENCE_PIPELINE_ON_STARTUP", "Optional", "Set to '1'/'true' to run the transcript intelligence pipeline on app startup. Disabled by default on HF to keep cold start fast."),
    ("AUTO_SYNC_TRANSCRIPTS_ON_STARTUP", "Optional", "Set to '1'/'true' to sync local transcript .txt files into the Excel workbook on startup. Disabled by default."),
    ("OPENAI_API_KEY",      "Required for Genie", "OpenAI API key for the Genie AI assistant (04_Genie.py)."),
    ("HF_TOKEN",            "Deployment", "HuggingFace token for pushing to the Space. Used in git remote: https://user:HF_TOKEN@huggingface.co/spaces/sebbo89/Earningscall2"),
]


@dataclass
class FnInfo:
    name: str
    lineno: int
    end_lineno: int
    refs: int
    calls: int


@dataclass
class ProbeInfo:
    name: str
    status: str
    seconds: float
    details: str


def _read(path: Path) -> str:
    return path.read_text(encoding="utf-8", errors="ignore")


def _git(cmd: list[str], default: str = "") -> str:
    try:
        out = subprocess.check_output(cmd, cwd=ROOT, text=True, stderr=subprocess.DEVNULL).strip()
        return out or default
    except Exception:
        return default


def _find_function_inventory(path: Path) -> list[FnInfo]:
    src = _read(path)
    try:
        mod = ast.parse(src)
    except Exception:
        return []
    call_counts: dict[str, int] = defaultdict(int)
    for node in ast.walk(mod):
        if not isinstance(node, ast.Call):
            continue
        fn = node.func
        if isinstance(fn, ast.Name):
            call_counts[fn.id] += 1
    out: list[FnInfo] = []
    for node in mod.body:
        if isinstance(node, (ast.FunctionDef, ast.AsyncFunctionDef)):
            refs = len(re.findall(r"\b" + re.escape(node.name) + r"\b", src))
            calls = call_counts.get(node.name, 0)
            out.append(FnInfo(node.name, node.lineno, getattr(node, "end_lineno", node.lineno), refs, calls))
    return out


def _count_pattern(path: Path, token: str) -> int:
    return _read(path).count(token)


def _extract_welcome_beats(path: Path) -> list[tuple[str, str, int]]:
    src = _read(path)
    try:
        mod = ast.parse(src)
    except Exception:
        return []
    out: list[tuple[str, str, int]] = []
    for node in ast.walk(mod):
        if not isinstance(node, ast.Call):
            continue
        if not isinstance(node.func, ast.Name) or node.func.id != "_section":
            continue
        if len(node.args) < 2:
            continue
        a0, a1 = node.args[0], node.args[1]
        if not isinstance(a0, ast.Constant) or not isinstance(a1, ast.Constant):
            continue
        if not isinstance(a0.value, str) or not isinstance(a1.value, str):
            continue
        out.append((a0.value, a1.value, getattr(node, "lineno", 0)))
    out.sort(key=lambda x: x[2])
    return out


def _extract_overview_areas(path: Path) -> list[dict]:
    src = _read(path)
    m = re.search(r"_OVERVIEW_AREA_CONFIG\s*=\s*\[(.*?)\]\n\n", src, flags=re.DOTALL)
    if not m:
        return []
    blob = m.group(1)
    areas: list[dict] = []
    for block in re.finditer(r"\{(.*?)\}", blob, flags=re.DOTALL):
        text = block.group(1)
        key = re.search(r'"key"\s*:\s*"([^"]+)"', text)
        title = re.search(r'"title"\s*:\s*"([^"]+)"', text)
        desc = re.search(r'"description"\s*:\s*"([^"]+)"', text)
        if key and title:
            areas.append({
                "key": key.group(1),
                "title": title.group(1),
                "description": desc.group(1) if desc else "",
            })
    return areas


def _extract_sync_steps(path: Path) -> list[str]:
    steps: list[str] = []
    for line in _read(path).splitlines():
        if "run_step(" in line and "scripts/" in line:
            steps.append(line.strip())
    return steps


def _grep_sheet_usage(sheet_name: str, files: list[Path]) -> list[str]:
    """Return list of files that mention the given sheet name."""
    hits: list[str] = []
    term = sheet_name.strip()
    for p in files:
        if not p.exists():
            continue
        if term in _read(p):
            hits.append(p.relative_to(ROOT).as_posix())
    return hits


def _section(title: str) -> list[str]:
    return [f"## {title}", ""]


def _md_table(headers: list[str], rows: Iterable[list[str]]) -> list[str]:
    lines = ["| " + " | ".join(headers) + " |", "|" + "|".join(["---"] * len(headers)) + "|"]
    for row in rows:
        lines.append("| " + " | ".join(str(c) for c in row) + " |")
    lines.append("")
    return lines


def _format_probe_payload(raw: str) -> str:
    value = str(raw or "").strip()
    if not value:
        return "-"
    try:
        obj = json.loads(value)
    except Exception:
        obj = None
    if isinstance(obj, dict):
        items: list[str] = []
        for key, val in obj.items():
            items.append(f"{key}={val}")
        return "; ".join(items) if items else "-"
    compact = " | ".join(part.strip() for part in value.splitlines() if part.strip())
    return compact[:220] + ("..." if len(compact) > 220 else "")


def _run_probe(name: str, code: str, timeout_seconds: int) -> ProbeInfo:
    started = time.perf_counter()
    try:
        out = subprocess.check_output(
            ["python3", "-c", code],
            cwd=ROOT,
            text=True,
            stderr=subprocess.STDOUT,
            timeout=timeout_seconds,
        )
        elapsed = time.perf_counter() - started
        lines = [line.strip() for line in out.splitlines() if line.strip()]
        payload = lines[-1] if lines else ""
        return ProbeInfo(name=name, status="ok", seconds=elapsed, details=_format_probe_payload(payload))
    except subprocess.TimeoutExpired as exc:
        elapsed = time.perf_counter() - started
        snippet = ""
        if exc.stdout:
            snippet = str(exc.stdout).strip().splitlines()[-1]
        return ProbeInfo(name=name, status="timeout", seconds=elapsed, details=_format_probe_payload(snippet or f"timed out after {timeout_seconds}s"))
    except subprocess.CalledProcessError as exc:
        elapsed = time.perf_counter() - started
        output = (exc.output or "").strip()
        last_line = output.splitlines()[-1] if output else f"exit={exc.returncode}"
        return ProbeInfo(name=name, status="error", seconds=elapsed, details=_format_probe_payload(last_line))
    except Exception as exc:
        elapsed = time.perf_counter() - started
        return ProbeInfo(name=name, status="error", seconds=elapsed, details=str(exc))


def _collect_runtime_probes() -> list[ProbeInfo]:
    probes: list[ProbeInfo] = []
    probes.append(_run_probe(
        "Workbook resolver",
        """
import json, sys
from pathlib import Path
sys.path.insert(0, str((Path.cwd() / "app").resolve()))
from utils.workbook_source import resolve_financial_data_xlsx
path = resolve_financial_data_xlsx([])
print(json.dumps({"resolved": bool(path), "path": path or ""}))
""".strip(), timeout_seconds=12,
    ))
    probes.append(_run_probe(
        "FinancialDataProcessor.load_data()",
        """
import json, sys
from pathlib import Path
sys.path.insert(0, str((Path.cwd() / "app").resolve()))
from data_processor import FinancialDataProcessor
dp = FinancialDataProcessor()
dp.load_data()
metrics_rows = int(len(dp.df_metrics.index)) if getattr(dp, "df_metrics", None) is not None else 0
segments_rows = int(len(dp.df_segments.index)) if getattr(dp, "df_segments", None) is not None else 0
print(json.dumps({"metrics_rows": metrics_rows, "segments_rows": segments_rows, "has_data_path": bool(dp.data_path)}))
""".strip(), timeout_seconds=28,
    ))
    probes.append(_run_probe(
        "SQLite intelligence tables",
        """
import json, sqlite3
from pathlib import Path
candidates = [Path("earningscall_intelligence.db"), Path("app/earningscall_intelligence.db")]
db_path = next((p for p in candidates if p.exists()), None)
if not db_path:
    print(json.dumps({"db_found": False}))
    raise SystemExit(0)
conn = sqlite3.connect(str(db_path))
tables = {row[0] for row in conn.execute("SELECT name FROM sqlite_master WHERE type='table'")}
counts = {}
for table in ("transcript_highlights", "transcript_topics", "topic_metrics"):
    if table in tables:
        counts[table] = int(conn.execute(f"SELECT COUNT(*) FROM {table}").fetchone()[0])
conn.close()
print(json.dumps({"db_found": True, "db_path": str(db_path), **counts}))
""".strip(), timeout_seconds=8,
    ))
    return probes


def _get_recent_git_log(n: int = 15) -> list[tuple[str, str, str]]:
    """Returns list of (short_sha, date, message)."""
    try:
        raw = subprocess.check_output(
            ["git", "log", f"-{n}", "--pretty=format:%h|%ad|%s", "--date=short"],
            cwd=ROOT, text=True, stderr=subprocess.DEVNULL,
        ).strip()
        rows = []
        for line in raw.splitlines():
            parts = line.split("|", 2)
            if len(parts) == 3:
                rows.append((parts[0], parts[1], parts[2]))
        return rows
    except Exception:
        return []


def _get_workbook_sheet_info() -> list[tuple[str, str, str]]:
    """Probe local Excel workbook and return (sheet_name, dimensions, first_header_cols)."""
    try:
        import openpyxl
        sys_path_hack = str(ROOT / "app")
        import sys
        sys.path.insert(0, sys_path_hack)
        from utils.workbook_source import resolve_financial_data_xlsx  # type: ignore
        path = resolve_financial_data_xlsx([])
        if not path:
            return []
        wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
        results = []
        for name in wb.sheetnames:
            ws = wb[name]
            rows = ws.max_row or 0
            cols = ws.max_column or 0
            headers = []
            for row in ws.iter_rows(min_row=1, max_row=1, values_only=True):
                headers = [str(c) for c in row if c is not None]
                break
            hdr_str = ", ".join(headers[:6])
            if len(headers) > 6:
                hdr_str += f" (+{len(headers) - 6} more)"
            results.append((name, f"{rows}r × {cols}c", hdr_str))
        wb.close()
        return results
    except Exception as exc:
        return [("(probe failed)", str(exc), "")]


def build() -> str:
    now = dt.datetime.now(dt.timezone.utc).strftime("%Y-%m-%d %H:%M:%S UTC")
    sha = _git(["git", "rev-parse", "--short", "HEAD"], "unknown")
    branch = _git(["git", "branch", "--show-current"], "unknown")

    lines: list[str] = []
    lines.append("# Developer Insights Bible — The Attention Economy")
    lines.append("## Complete Technical Handover Document")
    lines.append("")
    lines.append(f"**Generated:** {now}")
    lines.append(f"**Git branch:** `{branch}` — commit `{sha}`")
    lines.append(f"**Repository root:** `{ROOT}`")
    lines.append(f"**HuggingFace Space:** `https://sebbo89-earningscall2.hf.space/`")
    lines.append(f"**HF Space repo:** `https://huggingface.co/spaces/sebbo89/Earningscall2`")
    lines.append("")
    lines.append("> This document is auto-generated by `scripts/generate_current_bible.py`.")
    lines.append("> Regenerate at any time with: `python3 scripts/refresh_dev_manual.py`")
    lines.append("> **Must be run after every change to `app/`, `scripts/`, or `.streamlit/`.**")
    lines.append("")

    # ── TABLE OF CONTENTS ──────────────────────────────────────────────────────
    lines.append("## Table of Contents")
    lines.append("")
    toc_sections = [
        "1. Project Overview",
        "2. Architecture & Entry Points",
        "3. Page Inventory",
        "4. Home Page Render Order (Welcome.py)",
        "5. Overview Page Architecture",
        "6. Excel Workbook — Complete Sheet Inventory",
        "7. Data Flow: Excel → Processors → Pages",
        "8. Utils Module Catalog",
        "9. Transcript Intelligence Pipeline",
        "10. Environment Variables & Secrets",
        "11. Deployment — HuggingFace Spaces",
        "12. Function Inventory (Key Files)",
        "13. Welcome.py Helper Functions (Active vs Legacy)",
        "14. Chart & Component Footprint",
        "15. Known Operational Risks",
        "16. Runtime Diagnostics (Auto-Probed)",
        "17. Recent Git History",
        "18. Regeneration & Artifact Policy",
    ]
    for item in toc_sections:
        lines.append(f"- {item}")
    lines.append("")

    # ── 1. PROJECT OVERVIEW ────────────────────────────────────────────────────
    lines += _section("1) Project Overview")
    lines.append("**The Attention Economy** is a multi-page Streamlit dashboard that tracks")
    lines.append("14 major media/tech companies (Alphabet, Amazon, Apple, Meta, Microsoft,")
    lines.append("Netflix, Disney, Comcast, Spotify, Roku, Warner Bros. Discovery, Paramount,")
    lines.append("Snap, Pinterest) across revenue, segments, subscribers, stock performance,")
    lines.append("and macro advertising market data.")
    lines.append("")
    lines.append("**Primary data source:** A single Excel workbook (`Earnings + stocks copy.xlsx`)")
    lines.append("stored in `app/attached_assets/`. This file is ~17MB and tracked in git.")
    lines.append("The workbook contains 40 sheets covering financials, segments, macro data,")
    lines.append("country-level advertising, stock prices, and transcripts.")
    lines.append("")
    lines.append("**Secondary data source:** Google Sheets (same workbook exported as XLSX).")
    lines.append("Used as fallback only when no valid local file is detected.")
    lines.append("")
    lines.append("**Intelligence layer:** SQLite database (`earningscall_intelligence.db`)")
    lines.append("built from earnings call transcript .txt files via the intelligence pipeline.")
    lines.append("Powers the transcript quotes strip (Human Voice) and Genie AI chat.")
    lines.append("")

    # ── 2. ARCHITECTURE ────────────────────────────────────────────────────────
    lines += _section("2) Architecture & Entry Points")
    lines.append("```")
    lines.append("app.py                          ← Git root entrypoint")
    lines.append("  └─ sys.path.insert(0, 'app')  ← Makes app/ importable")
    lines.append("  └─ exec(app/Welcome.py)        ← Streamlit entry page")
    lines.append("")
    lines.append("app/")
    lines.append("  Welcome.py                    ← Home page (scrolling narrative)")
    lines.append("  pages/")
    lines.append("    00_Overview.py              ← Macro / global ad market")
    lines.append("    01_Earnings.py              ← Company deep-dive")
    lines.append("    02_Stocks.py                ← Stock charts")
    lines.append("    03_Editorial.py             ← AI-generated insights")
    lines.append("    04_Genie.py                 ← AI chat assistant")
    lines.append("  utils/                        ← Shared utilities (~40 modules)")
    lines.append("  attached_assets/              ← Excel workbook + logos")
    lines.append("  data_processor.py             ← Main FinancialDataProcessor class")
    lines.append("  .streamlit/config.toml        ← Streamlit config (no theme set)")
    lines.append("")
    lines.append("scripts/                        ← Pipeline scripts (run outside Streamlit)")
    lines.append("earningscall_transcripts/       ← Transcript .txt files + index CSVs")
    lines.append("earningscall_intelligence.db    ← SQLite DB (gitignored, built at runtime)")
    lines.append("reports/                        ← Auto-generated dev bible + assets")
    lines.append("```")
    lines.append("")
    lines.append("**Theme:** No Streamlit theme is set in config.toml — the app uses custom")
    lines.append("inline CSS to enforce a dark theme (`#0d1117` background) on the Home page.")
    lines.append("The Earnings and Overview pages use a semi-light theme via `utils/styles.py`")
    lines.append("with chart text tuned for dark background rendering.")
    lines.append("")

    # ── 3. PAGE INVENTORY ──────────────────────────────────────────────────────
    lines += _section("3) Page Inventory")
    page_rows = []
    for p in sorted((ROOT / "app" / "pages").glob("*.py")):
        name = p.name
        desc = PAGE_DESCRIPTIONS.get(name, "—")
        src = _read(p)
        page_rows.append([f"`{p.relative_to(ROOT)}`", str(len(src.splitlines())), desc])
    lines += _md_table(["File", "Lines", "Description"], page_rows)

    # ── 4. HOME PAGE RENDER ORDER ──────────────────────────────────────────────
    lines += _section("4) Home Page Render Order (Welcome.py)")
    beats = _extract_welcome_beats(ROOT / "app" / "Welcome.py")
    pre_items = [
        "Dark CSS + fonts injected (`st.markdown` with `<style>`)",
        "Excel workbook resolved (`resolve_financial_data_xlsx`)",
        "Intelligence pipeline freshness check (only if `AUTO_REFRESH...=1`)",
        "Hero section: KPI strip (Global Ad Spend, Tracked Revenue, Market Cap) + narrative",
        "`The World` — orthographic globe choropleth (ad spend % GDP per country)",
        "`The Structural Shift` — animated donut (TV→Internet ad share 1999–2024)",
        "`Attention + Duopoly` — animated scene (Google+Meta combined share)",
        "`Concentration` — bitcoin-style horizontal bar showing duopoly + human sidebars",
        "`Revenue Anatomy` — full-width flex bars per company (total + ad% breakdown)",
    ]
    for i, item in enumerate(pre_items, start=1):
        lines.append(f"{i}. {item}")
    offset = len(pre_items)
    for i, (label, headline, ln) in enumerate(beats, start=1 + offset):
        lines.append(f"{i}. `{label}` — *{headline}* (`app/Welcome.py:{ln}`)")
    lines.append(f"{len(pre_items)+len(beats)+1}. Market tape (stock strip, `_render_stock_price_strip`)")
    lines.append(f"{len(pre_items)+len(beats)+2}. Gateway section (→ Overview / Earnings / Genie)")
    lines.append("")

    # ── 5. OVERVIEW ARCHITECTURE ───────────────────────────────────────────────
    lines += _section("5) Overview Page Architecture (00_Overview.py)")
    areas = _extract_overview_areas(ROOT / "app" / "pages" / "00_Overview.py")
    lines.append("Navigator model: 8-section single-view with per-section `st.stop()` exits.")
    lines.append("Each section is self-contained — selecting a section renders only that section.")
    lines.append("")
    if areas:
        area_rows = [[f"`{a['key']}`", a["title"], a["description"]] for a in areas]
        lines += _md_table(["Key", "Title", "Description"], area_rows)
    else:
        lines.append("_(Could not extract `_OVERVIEW_AREA_CONFIG` — check 00_Overview.py)_")
        lines.append("")

    # ── 6. EXCEL WORKBOOK SHEET INVENTORY ─────────────────────────────────────
    lines += _section("6) Excel Workbook — Complete Sheet Inventory")
    lines.append("**File:** `app/attached_assets/Earnings + stocks  copy.xlsx` (~17MB, git-tracked)")
    lines.append("")
    lines.append("This single file is the primary data source for all pages.")
    lines.append("The `resolve_financial_data_xlsx()` function in `utils/workbook_source.py`")
    lines.append("auto-detects it and validates: ≥43 sheets, required tabs present,")
    lines.append("≥5 companies, ≥8 years, earliest year ≤ 2015.")
    lines.append("")

    # Static catalog (always accurate — maintained manually)
    lines.append("### Sheet Catalog (Purpose & Usage)")
    lines.append("")
    cat_rows = []
    for (name, purpose, key_cols, used_by) in SHEET_CATALOG:
        cat_rows.append([f"`{name}`", purpose, key_cols, used_by])
    lines += _md_table(["Sheet Name", "Purpose", "Key Columns", "Used By"], cat_rows)

    # Live probe (shows actual row/col counts)
    lines.append("### Live Sheet Dimensions (Probed at Generation Time)")
    lines.append("")
    sheet_info = _get_workbook_sheet_info()
    if sheet_info:
        lines += _md_table(["Sheet Name", "Dimensions", "First Columns (up to 6)"], sheet_info)
    else:
        lines.append("_(Workbook probe unavailable — run from repo root with `app/` in path)_")
        lines.append("")

    # ── 7. DATA FLOW ───────────────────────────────────────────────────────────
    lines += _section("7) Data Flow: Excel → Processors → Pages")
    lines.append("```")
    lines.append("Excel Workbook (app/attached_assets/*.xlsx)")
    lines.append("  │")
    lines.append("  ├─ resolve_financial_data_xlsx()        ← utils/workbook_source.py")
    lines.append("  │    Resolution order:")
    lines.append("  │    1. Explicit local candidates passed by caller")
    lines.append("  │    2. Auto-detect *.xlsx in app/attached_assets/")
    lines.append("  │    3. Download from GOOGLE_SHEET_URL (20s timeout)")
    lines.append("  │")
    lines.append("  ├─ FinancialDataProcessor               ← app/data_processor.py")
    lines.append("  │    .load_data()  →  reads sheets:")
    lines.append("  │      Company_metrics_earnings_values  → df_metrics")
    lines.append("  │      Company_yearly_segments_values   → df_segments")
    lines.append("  │      Company_subscribers_values       → df_subscribers")
    lines.append("  │      Company_Employees                → employee data")
    lines.append("  │      Company_revenue_by_region        → regional data")
    lines.append("  │    Cached via get_data_processor() in utils/state_management.py")
    lines.append("  │")
    lines.append("  ├─ _read_excel_sheet_cached()           ← Welcome.py internal")
    lines.append("  │    Reads individual sheets on demand (KPIs, map, anatomy, etc.)")
    lines.append("  │    Cached with @st.cache_data(ttl=4h)")
    lines.append("  │")
    lines.append("  ├─ load_quarterly_segments()            ← 01_Earnings.py internal")
    lines.append("  │    Reads per-company 'X Quarterly Segments' sheets")
    lines.append("  │")
    lines.append("  └─ load_combined_stock_market_data()    ← utils/workbook_market_data.py")
    lines.append("       Reads 'Stocks & Crypto' sheet")
    lines.append("       → Used by Welcome.py (market tape) and 02_Stocks.py")
    lines.append("")
    lines.append("SQLite (earningscall_intelligence.db)")
    lines.append("  Built by: scripts/build_intelligence_db.py")
    lines.append("  Tables: transcript_highlights, transcript_topics, topic_metrics")
    lines.append("  Used by: Welcome.py (Human Voice strip), 04_Genie.py (Genie AI context)")
    lines.append("```")
    lines.append("")

    # ── 8. UTILS CATALOG ───────────────────────────────────────────────────────
    lines += _section("8) Utils Module Catalog (`app/utils/`)")
    utils_dir = ROOT / "app" / "utils"
    util_rows = []
    for fname, desc in UTILS_CATALOG.items():
        p = utils_dir / fname
        exists = "✓" if p.exists() else "✗ missing"
        lines_count = len(_read(p).splitlines()) if p.exists() else 0
        util_rows.append([f"`{fname}`", exists, str(lines_count) if lines_count else "—", desc])
    lines += _md_table(["Module", "Exists", "Lines", "Description"], util_rows)

    # Also list any utils not in the catalog
    all_utils = sorted(f.name for f in utils_dir.glob("*.py") if f.name != "__init__.py")
    undocumented = [f for f in all_utils if f not in UTILS_CATALOG]
    if undocumented:
        lines.append("**Additional utils (not yet catalogued):**")
        for f in undocumented:
            p = utils_dir / f
            n = len(_read(p).splitlines()) if p.exists() else 0
            lines.append(f"- `{f}` ({n} lines)")
        lines.append("")

    # ── 9. TRANSCRIPT INTELLIGENCE PIPELINE ───────────────────────────────────
    lines += _section("9) Transcript Intelligence Pipeline")
    lines.append("**Orchestrator:** `scripts/sync_all_intelligence.py`")
    lines.append("")
    lines.append("**Input:** Earnings call transcripts as `.txt` files in `earningscall_transcripts/`")
    lines.append("**Output:** SQLite DB + CSV files consumed by Home page and Genie AI")
    lines.append("")
    steps = _extract_sync_steps(ROOT / "scripts" / "sync_all_intelligence.py")
    if steps:
        for s in steps:
            lines.append(f"- `{s}`")
    else:
        lines.append("- _(pipeline steps not parsed — check sync_all_intelligence.py)_")
    lines.append("")
    lines.append("**Transcript CSVs in `earningscall_transcripts/`:**")
    csv_files = [
        ("transcript_index.csv",       "Master index: company/year/quarter/file path"),
        ("transcript_topics.csv",      "Topics extracted per transcript"),
        ("transcript_kpis.csv",        "KPI values mentioned in transcripts"),
        ("transcript_highlights.csv",  "Key quote highlights per transcript"),
        ("overview_iconic_quotes.csv", "Hand-picked / top-ranked quotes for Overview"),
        ("topic_metrics.csv",          "Aggregated topic frequency metrics"),
        ("generated_insights_latest.csv", "AI-generated insight text (latest run)"),
    ]
    for csv_name, desc in csv_files:
        exists = (ROOT / "earningscall_transcripts" / csv_name).exists()
        status = "present" if exists else "**MISSING**"
        lines.append(f"- `{csv_name}` ({status}): {desc}")
    lines.append("")
    lines.append("**Trigger:** Pipeline only runs automatically if `AUTO_REFRESH_INTELLIGENCE_PIPELINE_ON_STARTUP=1`.")
    lines.append("Otherwise run manually: `python3 scripts/sync_all_intelligence.py`")
    lines.append("")

    # ── 10. ENVIRONMENT VARIABLES ─────────────────────────────────────────────
    lines += _section("10) Environment Variables & Secrets")
    env_rows = [[name, req, desc] for name, req, desc in ENV_VARS]
    lines += _md_table(["Variable", "Required?", "Description"], env_rows)
    lines.append("**Where to set on HuggingFace:** Space → Settings → Repository secrets")
    lines.append("")

    # ── 11. DEPLOYMENT ────────────────────────────────────────────────────────
    lines += _section("11) Deployment — HuggingFace Spaces")
    lines.append("**Platform:** HuggingFace Spaces (free tier, Streamlit runtime)")
    lines.append(f"**Live URL:** `https://sebbo89-earningscall2.hf.space/`")
    lines.append(f"**Space repo:** `https://huggingface.co/spaces/sebbo89/Earningscall2`")
    lines.append("")
    lines.append("**Git remote setup:**")
    lines.append("```bash")
    lines.append("git remote add hf https://huggingface.co/spaces/sebbo89/Earningscall2")
    lines.append("git push hf main")
    lines.append("```")
    lines.append("")
    lines.append("**Push workflow (from CODEX_RULES.md):**")
    lines.append("```bash")
    lines.append("# 1. Compile-check all 3 pages")
    lines.append("PYTHONPYCACHEPREFIX=/tmp/pycache python3 -m py_compile app/Welcome.py")
    lines.append("PYTHONPYCACHEPREFIX=/tmp/pycache python3 -m py_compile app/pages/00_Overview.py")
    lines.append("PYTHONPYCACHEPREFIX=/tmp/pycache python3 -m py_compile app/pages/01_Earnings.py")
    lines.append("# 2. Regenerate dev bible")
    lines.append("python3 scripts/refresh_dev_manual.py")
    lines.append("# 3. Push to HuggingFace")
    lines.append("git push hf main")
    lines.append("```")
    lines.append("")
    lines.append("**Cold start behaviour:** Free HF tier sleeps containers after inactivity.")
    lines.append("Container wake takes ~30–60s (unavoidable). Data loading is ~2s (local XLSX).")
    lines.append("")
    lines.append("**Large files — NEVER commit:**")
    lines.append("- `earningscall_intelligence.db` (gitignored)")
    lines.append("- `app/attached_assets/HeroVideo.mp4` (gitignored)")
    lines.append("- Any file over 5MB (except the tracked XLSX which is ~17MB)")
    lines.append("")

    # ── 12. FUNCTION INVENTORY ────────────────────────────────────────────────
    lines += _section("12) Function Inventory (Key Files)")
    inv_rows = []
    for rel in KEY_FILES:
        p = ROOT / rel
        if not p.exists():
            inv_rows.append([f"`{rel}`", "**MISSING**", "—", "—"])
            continue
        funcs = _find_function_inventory(p)
        src = _read(p)
        classes = len(re.findall(r"^class\s+", src, flags=re.MULTILINE))
        inv_rows.append([f"`{rel}`", str(len(src.splitlines())), str(len(funcs)), str(classes)])
    lines += _md_table(["File", "Lines", "Functions", "Classes"], inv_rows)

    # ── 13. WELCOME.PY ACTIVE VS LEGACY HELPERS ───────────────────────────────
    lines += _section("13) Welcome.py Helper Functions (Active vs Legacy)")
    welcome_funcs = _find_function_inventory(ROOT / "app" / "Welcome.py")
    active_rows, legacy_rows = [], []
    for fn in welcome_funcs:
        row = [f"`{fn.name}`", str(fn.lineno), str(fn.calls), str(fn.refs)]
        (active_rows if fn.calls > 0 else legacy_rows).append(row)
    lines.append("**Active** (called at least once in file):")
    lines += _md_table(["Function", "Line", "Internal Calls", "Name References"], active_rows)
    lines.append("**Possibly legacy** (call count = 0 within Welcome.py):")
    lines += _md_table(["Function", "Line", "Internal Calls", "Name References"], legacy_rows)

    # ── 14. CHART & COMPONENT FOOTPRINT ──────────────────────────────────────
    lines += _section("14) Chart & Component Footprint")
    chart_rows = []
    for rel in ["app/Welcome.py", "app/pages/00_Overview.py", "app/pages/01_Earnings.py", "app/pages/04_Genie.py"]:
        p = ROOT / rel
        if not p.exists():
            continue
        chart_rows.append([
            f"`{rel}`",
            str(_count_pattern(p, "st.plotly_chart(")),
            str(_count_pattern(p, "st.components.v1.html(")),
            str(_count_pattern(p, "@st.cache_data")),
            str(_count_pattern(p, "st.stop(")),
        ])
    lines += _md_table(["File", "Plotly Calls", "HTML Components", "@st.cache_data", "st.stop Calls"], chart_rows)

    # ── 15. KNOWN OPERATIONAL RISKS ───────────────────────────────────────────
    lines += _section("15) Known Operational Risks")
    risks = [
        ("CSS bleeding (global selectors)",
         "Welcome.py injects CSS using `.stApp`, `body`, `.main` — these apply globally. "
         "Earnings.py charts now use `render_plotly(light_theme=False)` by default (dark theme). "
         "If Earnings text appears invisible, check `render_plotly()` in 01_Earnings.py."),
        ("M2 sheet dual naming",
         "Old workbooks have sheet `M2_values` with columns `USD observation_date`/`WM2NS`. "
         "New workbooks use sheet `M2` with columns `observation_date`/`M2SL`. "
         "All 4 load points now try both names via try-loop."),
        ("Overview complexity",
         "`00_Overview.py` is >10k lines with many branch `st.stop()` exits — hard to reason, easy to regress."),
        ("HF cold start latency",
         "Free HF tier: ~30–60s container wake (unavoidable). Data load itself is ~2s via local XLSX."),
        ("SQLite DB not in git",
         "`earningscall_intelligence.db` is gitignored. Genie AI and Human Voice strip fail silently if DB missing."),
        ("No auto-regeneration of bible",
         "Bible does NOT auto-regenerate. Must run `python3 scripts/refresh_dev_manual.py` manually after every change to app/, scripts/, or .streamlit/."),
        ("Unused utility drift",
         "`app/utils/` has ~40 modules; several may be legacy. `insights_loader_fixed.py`, `optimized_data_loader.py` look like duplicates."),
    ]
    for title, desc in risks:
        lines.append(f"- **{title}:** {desc}")
    lines.append("")

    # ── 16. RUNTIME DIAGNOSTICS ───────────────────────────────────────────────
    lines += _section("16) Runtime Diagnostics (Auto-Probed at Generation Time)")
    probe_rows = []
    probes = _collect_runtime_probes()
    for probe in probes:
        probe_rows.append([f"`{probe.name}`", probe.status, f"{probe.seconds:.2f}s", probe.details or "—"])
    lines += _md_table(["Check", "Status", "Time", "Details"], probe_rows)
    ok = [p for p in probes if p.status == "ok"]
    if ok:
        slowest = max(ok, key=lambda x: x.seconds)
        lines.append(f"- Slowest successful probe: **{slowest.name}** at `{slowest.seconds:.2f}s`")
    blocked = [p for p in probes if p.status in {"timeout", "error"}]
    if blocked:
        lines.append("- Checks needing attention:")
        for p in blocked:
            lines.append(f"  - `{p.name}` → `{p.status}` ({p.details})")
    lines.append("")

    # ── 17. RECENT GIT HISTORY ────────────────────────────────────────────────
    lines += _section("17) Recent Git History (Last 15 Commits)")
    git_rows = _get_recent_git_log(15)
    if git_rows:
        lines += _md_table(["SHA", "Date", "Commit Message"], git_rows)
    else:
        lines.append("_(git log unavailable)_")
        lines.append("")

    # ── 18. REGENERATION & ARTIFACT POLICY ───────────────────────────────────
    lines += _section("18) Regeneration & Artifact Policy")
    lines.append("**To regenerate this bible after any code change:**")
    lines.append("```bash")
    lines.append("python3 scripts/refresh_dev_manual.py")
    lines.append("```")
    lines.append("This runs:")
    lines.append("1. `scripts/generate_dev_manual_assets.py` — PNG charts")
    lines.append("2. `scripts/generate_current_bible.py` — this markdown file")
    lines.append("3. `scripts/generate_current_bible_pdf.py` — PDF export")
    lines.append("4. Removes legacy duplicate files")
    lines.append("5. Writes `reports/DEV_MANUAL_STATUS.json`")
    lines.append("")
    lines.append("**Canonical outputs (never recreate old filenames):**")
    lines.append("- `reports/Developer_Insights_Bible_CURRENT.md`")
    lines.append("- `reports/Developer_Insights_Bible_CURRENT_Full.pdf`")
    lines.append("")
    lines.append("**Legacy files that must NOT be recreated:**")
    lines.append("- `reports/Developer_Insights_Bible.md`")
    lines.append("- `reports/Developer_Insights_Bible_Full.pdf`")
    lines.append("")

    return "\n".join(lines).rstrip() + "\n"


if __name__ == "__main__":
    OUT.parent.mkdir(parents=True, exist_ok=True)
    OUT.write_text(build(), encoding="utf-8")
    print(f"Wrote {OUT}")
